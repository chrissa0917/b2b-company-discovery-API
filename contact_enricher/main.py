from __future__ import annotations

import asyncio
import csv
import hmac
import json
import os
import re
import time
import uuid
from pathlib import Path

from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, RedirectResponse
from openpyxl import Workbook, load_workbook

from .verified_enricher import enrich_rows

BASE = Path(__file__).resolve().parent.parent
DATA = Path(os.getenv("DATA_DIR", BASE / "data"))
DATA.mkdir(parents=True, exist_ok=True)
JOBS: dict[str, dict] = {}
API_KEY = os.getenv("CONTACT_ENRICHER_API_KEY", "")
DEFAULT_POSITIONS = [
    "Marketing Director",
    "Head of Marketing",
    "Marketing Manager",
    "Partnerships Manager",
    "Communications Director",
    "PR Manager",
    "Business Development Manager",
    "Founder",
    "CEO",
]

COMPANY_ALIASES = {
    "company",
    "company name",
    "companyname",
    "business",
    "business name",
    "organization",
    "organisation",
}
WEBSITE_ALIASES = {
    "website url",
    "website",
    "websiteurl",
    "company website",
    "companywebsite",
    "domain",
    "url",
}

app = FastAPI(title="Chrissa Automates Contact Enricher", version="2.2.0")


def require_api_key(request: Request) -> None:
    supplied = request.headers.get("x-contact-enricher-key", "")
    if not API_KEY or not supplied or not hmac.compare_digest(supplied, API_KEY):
        raise HTTPException(401, "Unauthorized")


def require_job_owner(request: Request, job: dict) -> None:
    supplied = request.headers.get("x-contact-enricher-user", "")
    owner = str(job.get("owner_id") or "")
    if not owner or not supplied or not hmac.compare_digest(supplied, owner):
        raise HTTPException(403, "This job belongs to a different account.")


def parse_positions(raw: str, mode: str) -> list[str]:
    if mode == "basic":
        return []
    raw = (raw or "").strip()
    if not raw:
        return DEFAULT_POSITIONS[:5]
    try:
        value = json.loads(raw)
        if isinstance(value, list):
            positions = [str(item).strip() for item in value if str(item).strip()]
            return list(dict.fromkeys(positions))[:8] or DEFAULT_POSITIONS[:5]
    except Exception:
        pass
    positions = [item.strip() for item in raw.replace("\r", "").replace("\n", ",").split(",") if item.strip()]
    return list(dict.fromkeys(positions))[:8] or DEFAULT_POSITIONS[:5]


def normalize_header(value: str) -> str:
    value = str(value or "").strip().lower()
    value = re.sub(r"[_\-]+", " ", value)
    value = re.sub(r"[^a-z0-9 ]+", "", value)
    return re.sub(r"\s+", " ", value).strip()


def read_headers(path: Path) -> list[str]:
    if path.suffix.lower() == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f)
            return [str(x or "").strip() for x in next(reader, [])]
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        first = next(ws.iter_rows(values_only=True), ())
        return [str(x or "").strip() for x in first]
    return []


def resolve_columns(headers: list[str]) -> tuple[str, str]:
    normalized = {normalize_header(header): header for header in headers if str(header or "").strip()}

    company_col = next((normalized[key] for key in COMPANY_ALIASES if key in normalized), "")
    website_col = next((normalized[key] for key in WEBSITE_ALIASES if key in normalized), "")

    if not company_col or not website_col:
        raise HTTPException(
            400,
            'We could not find your company and website columns. Use names like "Company" / "Website URL" or "company_name" / "website_url". Extra columns are okay.',
        )
    return company_col, website_col


def read_raw_input(path: Path) -> list[dict]:
    if path.suffix.lower() == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            return list(csv.DictReader(f))
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(x or "").strip() for x in rows[0]]
        return [dict(zip(headers, ["" if x is None else x for x in row])) for row in rows[1:]]
    raise ValueError("Only CSV and XLSX files are supported.")


def read_input(path: Path) -> list[dict]:
    headers = read_headers(path)
    company_col, website_col = resolve_columns(headers)
    raw_rows = read_raw_input(path)
    rows: list[dict] = []
    for raw in raw_rows:
        company = str(raw.get(company_col, "") or "").strip()
        website = str(raw.get(website_col, "") or "").strip()
        if company and website:
            rows.append({"Company": company, "Website URL": website})
    return rows


def write_outputs(job_id: str, rows: list[dict]) -> tuple[Path, Path]:
    csv_path = DATA / f"{job_id}-enriched.csv"
    xlsx_path = DATA / f"{job_id}-enriched.xlsx"
    fields: list[str] = []
    for row in rows:
        for key in row.keys():
            if key not in fields:
                fields.append(key)
    with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)
    wb = Workbook()
    ws = wb.active
    ws.title = "Enriched Contacts"
    ws.append(fields)
    for row in rows:
        ws.append([row.get(key, "") for key in fields])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(xlsx_path)
    return csv_path, xlsx_path


def progress_numbers(job: dict) -> tuple[int, int]:
    completed = int(job.get("completed", 0) or 0)
    total = int(job.get("total", 0) or 0)
    return completed, total


def estimate_job_seconds(total: int, mode: str) -> int:
    # A deliberately conservative first estimate. Once companies finish,
    # the live ETA below replaces this with the job's observed speed.
    seconds_per_company = 8 if mode == "targeted" else 5
    return max(30, int(total * seconds_per_company))


def timing_payload(job: dict) -> dict:
    completed, total = progress_numbers(job)
    started_at = float(job.get("started_at", 0) or 0)
    mode = str(job.get("mode") or "targeted")
    elapsed = max(0, int(time.time() - started_at)) if started_at else 0

    if completed > 0 and elapsed > 0 and total > completed:
        observed_seconds_per_company = elapsed / completed
        observed_seconds_per_company = min(max(observed_seconds_per_company, 2.0), 90.0)
        eta = int(observed_seconds_per_company * (total - completed))
    elif total > completed:
        initial_total = estimate_job_seconds(total, mode)
        eta = max(0, initial_total - elapsed)
    else:
        eta = 0

    return {
        "elapsed_seconds": elapsed,
        "eta_seconds": eta,
        "percent": round((completed / total) * 100) if total else 0,
    }


async def run_job(
    job_id: str,
    input_path: Path,
    concurrency: int,
    requested_positions: list[str],
    max_pages: int,
    mode: str,
):
    try:
        rows = read_input(input_path)
        JOBS[job_id].update(status="running", total=len(rows), completed=0, started_at=time.time())

        def progress(done, total):
            JOBS[job_id].update(completed=done, total=total)

        results = await enrich_rows(
            rows,
            requested_positions=requested_positions,
            concurrency=concurrency,
            use_search=mode == "targeted",
            max_pages=max_pages,
            deep_verify=True,
            progress_cb=progress,
        )
        csv_path, xlsx_path = write_outputs(job_id, results)
        JOBS[job_id].update(status="complete", completed=len(rows), csv=str(csv_path), xlsx=str(xlsx_path))
    except Exception as exc:
        JOBS[job_id].update(status="failed", error=str(exc))


@app.get("/health")
def health():
    return {"ok": True, "service": "chrissa-automates-contact-enricher"}


@app.get("/")
def home():
    return RedirectResponse("https://chrissaautomates.com/contact-enricher", status_code=302)


@app.post("/jobs/stage")
async def stage_job(
    request: Request,
    file: UploadFile = File(...),
    owner_id: str = Form(...),
    positions: str = Form(""),
    mode: str = Form("targeted"),
    concurrency: int = Form(4),
    max_pages: int = Form(12),
):
    require_api_key(request)
    owner_id = owner_id.strip()
    mode = mode.strip().lower()
    if mode not in {"basic", "targeted"}:
        raise HTTPException(400, "Invalid lookup mode.")
    if not owner_id:
        raise HTTPException(400, "Missing account owner.")
    suffix = Path(file.filename or "input.csv").suffix.lower()
    if suffix not in {".csv", ".xlsx", ".xlsm"}:
        raise HTTPException(400, "Upload a CSV or XLSX file.")

    job_id = uuid.uuid4().hex[:12]
    input_path = DATA / f"{job_id}-input{suffix}"
    input_path.write_bytes(await file.read())
    rows = read_input(input_path)
    if not rows:
        raise HTTPException(400, "We found the right columns, but no rows had both a company name and website.")
    if len(rows) > 2000:
        raise HTTPException(400, "Please upload 2,000 companies or fewer per job.")

    requested_positions = parse_positions(positions, mode)
    JOBS[job_id] = {
        "id": job_id,
        "owner_id": owner_id,
        "status": "staged",
        "completed": 0,
        "total": len(rows),
        "input_path": str(input_path),
        "positions": requested_positions,
        "mode": mode,
        "concurrency": min(max(concurrency, 1), 8),
        "max_pages": min(max(max_pages, 3), 20),
        "started_at": 0,
    }
    return {
        "id": job_id,
        "status": "staged",
        "total": len(rows),
        "positions": requested_positions,
        "mode": mode,
        "estimated_seconds": estimate_job_seconds(len(rows), mode),
    }


@app.post("/jobs/{job_id}/start")
async def start_job(request: Request, job_id: str):
    require_api_key(request)
    job = JOBS.get(job_id)
    if not job:
        raise HTTPException(404, "Job not found")
    require_job_owner(request, job)
    if job.get("status") not in {"staged", "failed"}:
        return {"id": job_id, "status": job.get("status"), "total": job.get("total", 0)}
    input_path = Path(job["input_path"])
    job["status"] = "queued"
    job["started_at"] = time.time()
    asyncio.create_task(
        run_job(
            job_id,
            input_path,
            int(job.get("concurrency", 4)),
            list(job.get("positions") or []),
            int(job.get("max_pages", 12)),
            str(job.get("mode") or "targeted"),
        )
    )
    return {
        "id": job_id,
        "status": "queued",
        "total": job.get("total", 0),
        "mode": job.get("mode"),
        "estimated_seconds": estimate_job_seconds(int(job.get("total", 0) or 0), str(job.get("mode") or "targeted")),
    }


@app.get("/jobs/{job_id}")
def job_status(request: Request, job_id: str):
    require_api_key(request)
    job = JOBS.get(job_id)
    if not job:
        raise HTTPException(404, "Job not found. Jobs are kept in memory until the service restarts.")
    require_job_owner(request, job)
    return {
        "id": job_id,
        "status": job.get("status"),
        "completed": job.get("completed", 0),
        "total": job.get("total", 0),
        "mode": job.get("mode", "targeted"),
        "error": job.get("error", ""),
        **timing_payload(job),
    }


@app.get("/jobs/{job_id}/download/{kind}")
def download(request: Request, job_id: str, kind: str):
    require_api_key(request)
    job = JOBS.get(job_id)
    if not job or job.get("status") != "complete":
        raise HTTPException(404, "Output is not ready.")
    require_job_owner(request, job)
    if kind not in {"csv", "xlsx"}:
        raise HTTPException(400, "Choose csv or xlsx.")
    path = job.get(kind)
    return FileResponse(path, filename=f"chrissa-automates-contact-enricher-{job_id}.{kind}")
